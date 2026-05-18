"""
Lógica de conciliação de clientes LLE
Contábil (Clientes) × Financeiro (Data Base)
"""
import unicodedata
import pandas as pd


# ── Listas canônicas do PROMPT ────────────────────────────────────────────────

DESCROPER_VALIDAS = [
    "importacao cheque receita",
    "pagamentos extemporaneos",
    "venda",
    "venda (gold)",
    "venda (importacao xml)",
    "venda (pisa)",
    "venda cupom fiscal",
    "venda cupom fiscal gold",
    "complemento icms st - saida",
    "complemento ipi venda",
    "bonificacao a clientes",
]

TIPTIT_VALIDOS = [
    "adiantamento",
    "antigo cred c6 pay 10x",
    "antigo cred c6 pay 12x",
    "antigo stone credito a vista",
    "antigo cartao credito parcelada",
    "boleto",
    "boleto registrado",
    "boleto registrado liquidado",
    "boleto registrado alterado",
    "boleto rejeitado",
    "boleto retorno/titulo vencido",
    "credito manual",
    "credito automatico",
    "deposito bancario",
    "dinheiro",
    "duplicata",
    "pix",
    "pix qr code presencial",
    "edi-dda",
]

# Fragmentos de TIPTIT que devem ser EXCLUÍDOS (cartões)
CARTOES_EXCLUIDOS = [
    "getnet tef",
    "cred parc",
    "credito a distancia",
    "credito a vista",
    "debito getnet",
    "debito- vis",
    "debito- mas",
    "debito- elo",
    "cred tef",
    "deb tef",
]

# CODPARC do parceiro SEPM — entra independente do TIPTIT
CODPARC_SEPM = 41007


# ── Normalização ──────────────────────────────────────────────────────────────

def norm(valor):
    """Lowercase + strip acentos para comparação case/accent-insensitive."""
    if pd.isna(valor):
        return ""
    s = str(valor).lower().strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s


def eh_cartao_excluido(tiptit_norm: str) -> bool:
    for frag in CARTOES_EXCLUIDOS:
        if frag in tiptit_norm:
            return True
    return False


# ── Leitura das bases ─────────────────────────────────────────────────────────

def ler_contabil(arquivo) -> pd.DataFrame:
    """Lê a base contábil (Clientes). Header na linha 1 (padrão)."""
    df = pd.read_excel(arquivo, dtype={"CODPARC": object})
    df.columns = [c.strip().upper() for c in df.columns]
    df["CODPARC"] = pd.to_numeric(df["CODPARC"], errors="coerce")
    df["VLRDESDOB"] = pd.to_numeric(df["VLRDESDOB"], errors="coerce").fillna(0)
    return df


def ler_financeiro(arquivo) -> pd.DataFrame:
    """Lê a base financeira (Data Base). Header sempre na linha 8 (índice 7)."""
    df = pd.read_excel(arquivo, header=7, dtype={"CODPARC": object})
    df.columns = [c.strip().upper() for c in df.columns]
    df["CODPARC"] = pd.to_numeric(df["CODPARC"], errors="coerce")
    df["VLRDESDOB"] = pd.to_numeric(df["VLRDESDOB"], errors="coerce").fillna(0)
    return df


# ── Filtro LLE na base financeira ─────────────────────────────────────────────

def filtrar_financeiro(df: pd.DataFrame) -> pd.DataFrame:
    """Aplica filtros canônicos LLE e retorna base filtrada."""
    df = df.copy()
    df["_descrnat"] = df["DESCRNAT"].apply(norm)
    df["_descroper"] = df["DESCROPER"].apply(norm)
    df["_tiptit"] = df["TIPTIT"].apply(norm)

    mask = (
        (df["_descrnat"] == "vendas notas fiscais")
        & (df["_descroper"].isin(DESCROPER_VALIDAS))
        & (
            (df["_tiptit"].isin(TIPTIT_VALIDOS))
            | (df["CODPARC"] == CODPARC_SEPM)
        )
        & (~df["_tiptit"].apply(eh_cartao_excluido))
    )

    return df[mask].drop(columns=["_descrnat", "_descroper", "_tiptit"])


# ── Detecção de órfãos ────────────────────────────────────────────────────────

def separar_orfaos(df_cli: pd.DataFrame, df_fin_filtrado: pd.DataFrame):
    """
    Retorna (df_cli_sem_orfaos, df_fin_sem_orfaos, orfaos_cli, orfaos_fin).
    Órfão = CODPARC nulo ou zero.
    """
    mask_cli = df_cli["CODPARC"].isna() | (df_cli["CODPARC"] == 0)
    mask_fin = df_fin_filtrado["CODPARC"].isna() | (df_fin_filtrado["CODPARC"] == 0)

    orfaos_cli = df_cli[mask_cli].copy()
    orfaos_fin = df_fin_filtrado[mask_fin].copy()

    df_cli_ok = df_cli[~mask_cli].copy()
    df_fin_ok = df_fin_filtrado[~mask_fin].copy()

    return df_cli_ok, df_fin_ok, orfaos_cli, orfaos_fin


def aplicar_atribuicoes_orfaos(
    df_cli: pd.DataFrame,
    orfaos_cli: pd.DataFrame,
    atribuicoes: dict,  # {index_original: codparc_novo}
) -> pd.DataFrame:
    """
    Aplica as atribuições manuais de CODPARC aos órfãos e devolve
    o dataframe contábil completo (sem órfãos pendentes).
    """
    orfaos_atualizados = orfaos_cli.copy()
    for idx, codparc in atribuicoes.items():
        orfaos_atualizados.loc[idx, "CODPARC"] = codparc

    # Remover os que ainda ficaram sem código (usuário deixou em branco)
    orfaos_atualizados = orfaos_atualizados[
        ~(orfaos_atualizados["CODPARC"].isna() | (orfaos_atualizados["CODPARC"] == 0))
    ]

    return pd.concat([df_cli, orfaos_atualizados], ignore_index=True)


# ── Conciliação ───────────────────────────────────────────────────────────────

def conciliar(df_cli: pd.DataFrame, df_fin: pd.DataFrame) -> pd.DataFrame:
    """
    Consolida por CODPARC e retorna tabela de diferenças.
    Somente parceiros com diferença != 0 (tolerância exata zero).
    """
    grp_cli = (
        df_cli.groupby("CODPARC")
        .agg(
            NOMEPARC=("NOMEPARC", "first"),
            QTD_CLI=("NUMNOTA", "count"),
            SOMA_CLI=("VLRDESDOB", "sum"),
        )
        .reset_index()
    )

    grp_fin = (
        df_fin.groupby("CODPARC")
        .agg(
            QTD_FIN=("NUMNOTA", "count"),
            SOMA_FIN=("VLRDESDOB", "sum"),
        )
        .reset_index()
    )

    merged = pd.merge(grp_cli, grp_fin, on="CODPARC", how="outer")

    # Preencher nulos
    merged["NOMEPARC"] = merged["NOMEPARC"].fillna(
        merged["CODPARC"].map(
            df_fin.dropna(subset=["CODPARC"])
            .groupby("CODPARC")["NOMEPARC"]
            .first()
        )
    )
    merged["QTD_CLI"] = merged["QTD_CLI"].fillna(0).astype(int)
    merged["SOMA_CLI"] = merged["SOMA_CLI"].fillna(0).round(2)
    merged["QTD_FIN"] = merged["QTD_FIN"].fillna(0).astype(int)
    merged["SOMA_FIN"] = merged["SOMA_FIN"].fillna(0).round(2)

    merged["DIFERENCA"] = (merged["SOMA_CLI"] - merged["SOMA_FIN"]).round(2)

    def status(row):
        if row["SOMA_FIN"] == 0 and row["SOMA_CLI"] != 0:
            return "Apenas no Contábil"
        if row["SOMA_CLI"] == 0 and row["SOMA_FIN"] != 0:
            return "Apenas no Financeiro"
        return "Diferença de valor"

    merged["STATUS"] = merged.apply(status, axis=1)

    # Apenas divergentes
    divergentes = merged[merged["DIFERENCA"] != 0].copy()

    # Ordenar por |Diferença| decrescente
    divergentes["_ABS"] = divergentes["DIFERENCA"].abs()
    divergentes = divergentes.sort_values("_ABS", ascending=False).drop(columns=["_ABS"])
    divergentes = divergentes.reset_index(drop=True)

    return divergentes


# ── Resumo macro ──────────────────────────────────────────────────────────────

def resumo_macro(df_cli, df_fin, df_dif, orfaos_cli, orfaos_fin):
    total_cli = round(df_cli["VLRDESDOB"].sum(), 2)
    total_fin = round(df_fin["VLRDESDOB"].sum(), 2)
    soma_orfaos = round(
        orfaos_cli["VLRDESDOB"].sum() + orfaos_fin["VLRDESDOB"].sum(), 2
    )
    diferenca_macro = round(total_cli - total_fin - soma_orfaos, 2)
    soma_parceiros = round(df_dif["DIFERENCA"].sum(), 2)
    valido = abs(diferenca_macro - soma_parceiros) <= 0.02

    return {
        "total_contabil": total_cli,
        "total_financeiro": total_fin,
        "soma_orfaos": soma_orfaos,
        "diferenca_macro": diferenca_macro,
        "soma_parceiros": soma_parceiros,
        "qtd_parceiros": len(df_dif),
        "status": "OK ✅" if valido else "DIVERGÊNCIA ⚠️",
    }


# ── Drill-down por CODPARC ────────────────────────────────────────────────────

def drill_down(codparc: int, df_cli: pd.DataFrame, df_fin: pd.DataFrame):
    """Retorna NFs do parceiro nas duas bases."""
    nfs_cli = df_cli[df_cli["CODPARC"] == codparc][
        ["NUMNOTA", "VLRDESDOB", "DTEMISSAO", "HISTORICO"]
    ].copy()
    nfs_cli = nfs_cli.rename(
        columns={
            "NUMNOTA": "NF",
            "VLRDESDOB": "Valor (R$)",
            "DTEMISSAO": "Data",
            "HISTORICO": "Histórico",
        }
    )

    nfs_fin = df_fin[df_fin["CODPARC"] == codparc][
        ["NUMNOTA", "VLRDESDOB", "DTEMISSAO", "DESCROPER"]
    ].copy()
    nfs_fin = nfs_fin.rename(
        columns={
            "NUMNOTA": "NF",
            "VLRDESDOB": "Valor (R$)",
            "DTEMISSAO": "Data",
            "DESCROPER": "Operação",
        }
    )

    # Resumo por NF
    grp_cli = (
        nfs_cli.groupby("NF")
        .agg(Σ_Contábil=("Valor (R$)", "sum"), Qtd_C=("Valor (R$)", "count"))
        .reset_index()
    )
    grp_fin = (
        nfs_fin.groupby("NF")
        .agg(Σ_Financeiro=("Valor (R$)", "sum"), Qtd_F=("Valor (R$)", "count"))
        .reset_index()
    )

    resumo_nf = pd.merge(grp_cli, grp_fin, on="NF", how="outer").fillna(0)
    resumo_nf["Σ_Contábil"] = resumo_nf["Σ_Contábil"].round(2)
    resumo_nf["Σ_Financeiro"] = resumo_nf["Σ_Financeiro"].round(2)
    resumo_nf["Δ"] = (resumo_nf["Σ_Contábil"] - resumo_nf["Σ_Financeiro"]).round(2)

    def status_nf(row):
        c, f = row["Σ_Contábil"], row["Σ_Financeiro"]
        delta = abs(row["Δ"])
        if c == 0 and f == 0:
            return "Compensa internamente"
        if f == 0 and c != 0:
            return "Só Contábil"
        if c == 0 and f != 0:
            return "Só Financeiro"
        if delta <= 0.02:
            return "OK"
        return "Diverge"

    resumo_nf["Status"] = resumo_nf.apply(status_nf, axis=1)
    resumo_nf = resumo_nf.sort_values("Δ", key=abs, ascending=False)

    return nfs_cli, nfs_fin, resumo_nf


# ── Export Excel ──────────────────────────────────────────────────────────────

def gerar_excel(df_filtrado, df_divergentes, resumo, orfaos_cli, orfaos_fin, observacoes=None) -> bytes:
    """Gera Excel com duas abas: Data base filtrado e Investigação Diferença."""
    import io
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl import Workbook

    wb = Workbook()

    # Cores LLE
    AZUL_ESCURO = "041747"
    AMARELO = "FAC318"
    VERDE = "0F8C3B"
    BRANCO = "FFFFFF"
    CINZA = "F2F2F2"
    BORDA_COR = "D9D9D9"

    def header_style(cell, bg=AZUL_ESCURO, fg=BRANCO):
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color=fg, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def borda_fina(cell):
        lado = Side(style="thin", color=BORDA_COR)
        cell.border = Border(left=lado, right=lado, top=lado, bottom=lado)

    # ── Aba única: Investigação Diferença ────────────────────────────────────
    ws2 = wb.active
    ws2.title = "Investigação Diferença"

    ws2["A1"] = "INVESTIGAÇÃO DIFERENÇA — Comparação Consolidada por Parceiro (CODPARC)"
    ws2["A1"].font = Font(bold=True, color=BRANCO, size=14, name="Calibri")
    ws2["A1"].fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    ws2.merge_cells("A1:H1")
    ws2["A1"].alignment = Alignment(horizontal="center")

    labels2 = [
        ("Diferença macro (Cont. − Fin.):", resumo["diferenca_macro"]),
        ("Soma diferenças por parceiro:", resumo["soma_parceiros"]),
        ("Qtd parceiros com diferença:", resumo["qtd_parceiros"]),
        ("Validação:", resumo["status"]),
    ]
    for i, (lb, vl) in enumerate(labels2, start=3):
        ws2[f"A{i}"] = lb
        ws2[f"A{i}"].font = Font(bold=True, name="Calibri")
        ws2[f"B{i}"] = vl
        if i == 6:  # Validação
            cor = VERDE if "OK" in str(vl) else "C00000"
            ws2[f"B{i}"].font = Font(bold=True, color=cor, name="Calibri")

    # Cabeçalho tabela
    cabecalhos = ["CODPARC", "NOMEPARC", "Qtd NFs Contábil", "Qtd NFs Financeiro", "Soma Contábil (R$)", "Soma Financeiro (R$)", "Diferença (R$)", "Status", "Observação do Analista"]
    for col_i, nome in enumerate(cabecalhos, start=1):
        cell = ws2.cell(row=10, column=col_i, value=nome)
        header_style(cell)
        borda_fina(cell)

    # Dados
    for row_i, (_, row_data) in enumerate(df_divergentes.iterrows(), start=11):
        obs = (observacoes or {}).get(int(row_data["CODPARC"]), "")
        vals = [
            row_data["CODPARC"], row_data["NOMEPARC"],
            row_data["QTD_CLI"], row_data["QTD_FIN"],
            row_data["SOMA_CLI"], row_data["SOMA_FIN"],
            row_data["DIFERENCA"], row_data["STATUS"], obs,
        ]
        for col_i, val in enumerate(vals, start=1):
            cell = ws2.cell(row=row_i, column=col_i, value=val)
            cell.font = Font(name="Calibri", size=10)
            fill_cor = BRANCO if (row_i - 11) % 2 == 0 else CINZA
            cell.fill = PatternFill("solid", fgColor=fill_cor)
            borda_fina(cell)
            if col_i in (5, 6, 7):  # Soma Contábil, Soma Financeiro, Diferença
                cell.number_format = '#,##0.00'

        # Cor da coluna observação
        obs_cell = ws2.cell(row=row_i, column=9)
        if obs_cell.value:
            obs_cell.fill = PatternFill("solid", fgColor="EAF4FF")
            obs_cell.font = Font(color="041747", name="Calibri", size=10)

        # Cor da coluna status
        status_cell = ws2.cell(row=row_i, column=8)
        s = row_data["STATUS"]
        if s == "Apenas no Contábil":
            status_cell.fill = PatternFill("solid", fgColor="FFE6E6")
            status_cell.font = Font(color="C00000", name="Calibri", size=10)
        elif s == "Apenas no Financeiro":
            status_cell.fill = PatternFill("solid", fgColor="FFE6E6")
            status_cell.font = Font(color="C00000", name="Calibri", size=10)
        else:
            status_cell.fill = PatternFill("solid", fgColor="FFF4CC")
            status_cell.font = Font(bold=True, color=AZUL_ESCURO, name="Calibri", size=10)

    ws2.freeze_panes = "A11"

    # Órfãos (se houver)
    ultima_linha = 11 + len(df_divergentes) + 2
    if len(orfaos_cli) > 0 or len(orfaos_fin) > 0:
        ws2.cell(row=ultima_linha, column=1, value="ÓRFÃOS SEM CODPARC").fill = \
            PatternFill("solid", fgColor=AMARELO)
        ws2.cell(row=ultima_linha, column=1).font = Font(bold=True, name="Calibri")
        ws2.merge_cells(f"A{ultima_linha}:F{ultima_linha}")

        sub_cab = ["Fonte", "NUMNOTA", "NOMEPARC", "Valor (R$)", "Data", "Observação"]
        for col_i, nome in enumerate(sub_cab, start=1):
            cell = ws2.cell(row=ultima_linha + 1, column=col_i, value=nome)
            header_style(cell, bg=AZUL_ESCURO)

        linha_orf = ultima_linha + 2
        for _, row_data in orfaos_cli.iterrows():
            ws2.cell(row=linha_orf, column=1, value="Contábil")
            ws2.cell(row=linha_orf, column=2, value=row_data.get("NUMNOTA", ""))
            ws2.cell(row=linha_orf, column=3, value=row_data.get("NOMEPARC", ""))
            ws2.cell(row=linha_orf, column=4, value=row_data.get("VLRDESDOB", 0))
            ws2.cell(row=linha_orf, column=5, value=row_data.get("DTEMISSAO", ""))
            linha_orf += 1
        for _, row_data in orfaos_fin.iterrows():
            ws2.cell(row=linha_orf, column=1, value="Financeiro")
            ws2.cell(row=linha_orf, column=2, value=row_data.get("NUMNOTA", ""))
            ws2.cell(row=linha_orf, column=3, value=row_data.get("NOMEPARC", ""))
            ws2.cell(row=linha_orf, column=4, value=row_data.get("VLRDESDOB", 0))
            ws2.cell(row=linha_orf, column=5, value=row_data.get("DTEMISSAO", ""))
            linha_orf += 1

    # Ajuste de largura das colunas
    for ws in [ws2]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
